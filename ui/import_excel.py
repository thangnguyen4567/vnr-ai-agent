import streamlit as st
import openpyxl
from langchain_core.documents import Document
from src.vectordb.vectordb import VectorDBManager
import tempfile
import os
from io import BytesIO
from typing import List, Dict, Any, Tuple


def filter_empty_columns(data: List[Dict[str, Any]], headers: List[str]) -> Tuple[List[str], List[Dict[str, Any]]]:
    """
    Lọc bỏ các cột rỗng (không có dữ liệu) từ dữ liệu Excel.

    Args:
        data: Danh sách các dictionary chứa dữ liệu
        headers: Danh sách tên các cột

    Returns:
        Tuple chứa (headers_filtered, data_filtered)
    """
    if not data or not headers:
        return headers, data

    # Kiểm tra từng cột xem có dữ liệu không
    valid_columns = []
    for header in headers:
        # Kiểm tra xem cột có ít nhất một giá trị không null không
        has_data = any(
            row.get(header) is not None and
            str(row.get(header)).strip() != ""
            for row in data
        )
        if has_data:
            valid_columns.append(header)

    # Lọc dữ liệu chỉ giữ lại các cột hợp lệ
    filtered_data = []
    for row in data:
        filtered_row = {col: row.get(col) for col in valid_columns}
        filtered_data.append(filtered_row)

    return valid_columns, filtered_data


def create_excel_template() -> bytes:
    """Đọc file Excel template có sẵn để download"""
    template_path = "static/example_template.xlsx"

    try:
        # Đọc file template có sẵn
        with open(template_path, "rb") as f:
            return f.read()
    except FileNotFoundError:
        # Fallback: tạo template đơn giản nếu không tìm thấy file
        st.warning(f"Không tìm thấy file template tại {template_path}. Tạo template mẫu...")
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Training_Data"

        # Headers cơ bản
        headers = ["content", "category", "source", "tags"]
        for col_idx, header in enumerate(headers, 1):
            worksheet.cell(row=1, column=col_idx, value=header)

        # Dữ liệu mẫu
        sample_data = [
            ["Nội dung tài liệu 1", "HR", "internal", "nhân sự"],
            ["Nội dung tài liệu 2", "Finance", "policy", "tài chính"]
        ]

        for row_idx, row_data in enumerate(sample_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                worksheet.cell(row=row_idx, column=col_idx, value=value)

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()


def import_excel():
    """Màn hình import dữ liệu từ file Excel"""
    st.title("📊 Import dữ liệu từ Excel")

    # Định nghĩa training options
    training_options = {
        "succession_data": "Training cho Agent kế nhiệm",
        "hrm_data": "Training cho Agent quản lý nhân sự",
    }

    # Section hướng dẫn và download template
    st.markdown("### 📋 Hướng dẫn Import")
    st.markdown("""
    1. **Tải template mẫu** để xem cấu trúc file Excel yêu cầu
    2. **Điền dữ liệu** vào file Excel theo định dạng mẫu
    3. **Upload file Excel** để import dữ liệu vào hệ thống
    """)

    try:
        template_data = create_excel_template()
        st.download_button(
            label="📥 Tải Template Excel",
            data=template_data,
            file_name="training_data_template.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="download_template",
            use_container_width=True
        )
    except Exception as e:
        st.error(f"Lỗi khi tải template: {str(e)}")

    selected_training_type = st.selectbox(
        "Loại dữ liệu training:",
        options=list(training_options.keys()),
        format_func=lambda x: training_options[x],
        help="Chọn loại dữ liệu training mà bạn muốn import"
    )

    # Lưu vào session state để sử dụng trong import
    st.session_state.selected_training_type = selected_training_type

    # Upload file Excel
    uploaded_file = st.file_uploader(
        "Chọn file Excel để import",
        type=["xlsx", "xls"],
        help="Chỉ chấp nhận file Excel (.xlsx, .xls)"
    )

    if uploaded_file is not None:
        # Hiển thị thông tin file
        st.success(f"Đã chọn file: {uploaded_file.name}")

        # Đọc file Excel
        try:
            # Đọc file Excel bằng openpyxl
            workbook = openpyxl.load_workbook(uploaded_file, data_only=True)

            # Hiển thị các sheet có sẵn
            sheet_names = workbook.sheetnames
            st.subheader("📋 Các sheet trong file")

            selected_sheet = st.selectbox(
                "Chọn sheet để import:",
                sheet_names,
                key="sheet_selector"
            )

            if selected_sheet:
                # Đọc dữ liệu từ sheet được chọn
                worksheet = workbook[selected_sheet]
                data = []

                # Lấy header từ dòng đầu tiên
                headers = []
                for col in range(1, worksheet.max_column + 1):
                    header_value = worksheet.cell(row=1, column=col).value
                    headers.append(header_value if header_value is not None else f"Column_{col}")

                # Đọc dữ liệu từ các dòng còn lại
                for row in range(2, worksheet.max_row + 1):
                    row_data = {}
                    for col_idx, header in enumerate(headers):
                        cell_value = worksheet.cell(row=row, column=col_idx + 1).value
                        row_data[header] = cell_value
                    data.append(row_data)

                # Lọc bỏ các cột rỗng
                original_column_count = len(headers)
                headers, data = filter_empty_columns(data, headers)
                filtered_column_count = len(headers)

                # Thông báo nếu có cột bị loại bỏ
                if original_column_count > filtered_column_count:
                    removed_columns = original_column_count - filtered_column_count
                    st.info(f"🧹 Đã loại bỏ {removed_columns} cột rỗng khỏi dữ liệu.")

                # Hiển thị preview dữ liệu
                st.subheader(f"👀 Preview dữ liệu từ sheet: {selected_sheet}")

                # Hiển thị thông tin tổng quan
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.metric("Số dòng", len(data))
                with col2:
                    st.metric("Số cột", len(headers))
                with col3:
                    # Đếm số ô có dữ liệu
                    data_count = sum(1 for row in data for value in row.values() if value is not None)
                    st.metric("Số ô có dữ liệu", data_count)

                # Hiển thị preview bảng (giới hạn 10 dòng đầu)
                preview_data = data[:10]
                if preview_data:
                    st.dataframe(preview_data, use_container_width=True)
                else:
                    st.info("Không có dữ liệu trong sheet này.")


                # Nút import dữ liệu
                if st.button("🚀 Import dữ liệu", type="primary", use_container_width=True):
                    # Lưu dữ liệu vào metadata
                    documents = []
                    for row in data:
                        content = row.get("content", "")
                        if 'content' in row:
                            del row['content']
                        doc = Document(page_content=content, metadata=row)
                        documents.append(doc)

                    print(documents)

                    # Tạo index_schema dựa trên các tên cột excel
                    index_schema = []
                    for header in headers:
                        index_schema.append({"name": header, "type": "text"})

                    vector_db = VectorDBManager()
                    training_type = st.session_state.selected_training_type
                    vector_db.add_documents(documents, training_type, index_schema)

                    st.success(f"✅ Đã import thành công !")

        except Exception as e:
            st.error(f"Lỗi khi đọc file Excel: {str(e)}")
            st.info("Vui lòng kiểm tra định dạng file Excel và thử lại.")



if __name__ == "__main__":
    import_excel()
